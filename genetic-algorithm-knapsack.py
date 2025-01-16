import random
from tkinter.font import Font
import matplotlib
import pygame
import matplotlib.pyplot as plt
from matplotlib.backends.backend_agg import FigureCanvasAgg
import openpyxl


matplotlib.use('Agg')  

def randomize_items_and_capacity(num_items, max_weight, max_value, max_capacity):
    items = [(random.randint(1, max_weight), random.randint(1, max_value)) for _ in range(num_items)]
    capacity = random.randint(1, max_capacity)
    return items, capacity


def fitness(individual, items, capacity):
    total_weight = 0
    total_value = 0
    for i in range(len(individual)):
        if individual[i] == 1:  
            total_weight += items[i][0]
            total_value += items[i][1]
    
    
    if total_weight > capacity:
        return 0
    else:
        return total_value

# Initialize population
def initialize_population(population_size, num_items):
    return [[random.choice([0, 1]) for _ in range(num_items)] for _ in range(population_size)]

# Tournament selection
def tournament_selection(population, items, capacity):
    tournament = random.sample(population, 3)  # Select 3 random individuals
    tournament.sort(key=lambda x: fitness(x, items, capacity), reverse=True)  # Sort by fitness value
    return tournament[0]  # Return the best individual

# 1-point crossover
def crossover(parent1, parent2):
    point = random.randint(1, len(parent1) - 1)
    child = parent1[:point] + parent2[point:]
    return child

# Mutation
def mutation(individual, mutation_rate=0.01):
    for i in range(len(individual)):
        if random.random() < mutation_rate:
            individual[i] = 1 - individual[i]  # Flip the value (0 to 1 or 1 to 0)
    return individual

from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import openpyxl

def save_to_excel(items, capacity, best_fitness, best_generation, all_items, filename="genetic_algorithm_results.xlsx"):
    wb = openpyxl.Workbook()
 
    ws1 = wb.active
    ws1.title = "Initial Setup"

    ws1["A1"] = "Knapsack Capacity"
    ws1["B1"] = capacity
    ws1["A1"].font = Font(bold=True)

    ws1.append(["Item", "Weight", "Value"])
    for i, (weight, value) in enumerate(all_items, start=1):
        ws1.append([f"Item {i}", weight, value])

    for col_num in range(1, 4):
        cell = ws1.cell(row=2, column=col_num)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws2 = wb.create_sheet(title="Genetic Algorithm Results")

    headers = ["Best Generation", "Best Fitness", "Knapsack Capacity", "Selected Items"]
    ws2.append(headers)

    selected_items = ", ".join([f"Item {i + 1} (W: {item[0]}, V: {item[1]})" for i, item in enumerate(items)])
    ws2.append([best_generation, best_fitness, capacity, selected_items])

    header_font = Font(bold=True)
    alignment = Alignment(horizontal="center", vertical="center")
    for col_num, header in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col_num)
        cell.font = header_font
        cell.alignment = alignment

    for ws in [ws1, ws2]:
        for col in ws.columns:
            max_length = max(len(str(cell.value)) for cell in col if cell.value)
            adjusted_width = max_length + 2
            ws.column_dimensions[get_column_letter(col[0].column)].width = adjusted_width

    wb.save(filename)
    print(f"Arquivo Excel salvo como '{filename}'")


# Genetic Algorithm
def genetic_algorithm(population_size, num_items, generations, mutation_rate, items, capacity, screen):
    population = initialize_population(population_size, num_items)
    best_fitness_values = [] 

    x_vals, y_vals = [], []  
    best_overall_fitness = 0
    best_overall_individual = None
    best_generation = 0
    gen_count = 0  


    print("Initial items (weight, value):")
    for i, item in enumerate(items):
        print(f"Item {i + 1}: Weight = {item[0]}, Value = {item[1]}")
    print(f"Knapsack capacity: {capacity}")
    print("-" * 50)  # Divider line for clarity

    while gen_count < generations: 
        gen_count += 1
        new_population = []
        while len(new_population) < population_size:
            parent1 = tournament_selection(population, items, capacity)
            parent2 = tournament_selection(population, items, capacity)
            child = crossover(parent1, parent2)
            child = mutation(child, mutation_rate)
            new_population.append(child)
        population = new_population

        best_individual = max(population, key=lambda x: fitness(x, items, capacity))
        best_fitness = fitness(best_individual, items, capacity)

        if best_fitness > best_overall_fitness:
            best_overall_fitness = best_fitness
            best_overall_individual = best_individual
            best_generation = gen_count

        best_fitness_values.append(best_fitness) 
        x_vals.append(gen_count)
        y_vals.append(best_fitness)

        print(f"gen-{gen_count}, fitness-{best_fitness}")

        draw_plot(screen, x_vals, y_vals, generation=gen_count, fitness=best_fitness)

    best_items = [items[i] for i in range(len(items)) if best_overall_individual[i] == 1]
    save_to_excel(best_items, capacity, best_overall_fitness, best_generation, items)


# Draw the plot using Pygame
def draw_plot(screen: pygame.Surface, x: list, y: list, generation: int, fitness: int) -> None:
    """
    Draw a plot on a Pygame screen using Matplotlib.

    Parameters:
    - screen (pygame.Surface): The Pygame surface to draw the plot on.
    - x (list): The x-axis values.
    - y (list): The y-axis values.
    - generation (int): Current generation number.
    - fitness (int): Current fitness value.
    """
    fig, ax = plt.subplots(figsize=(5, 5), dpi=100)
    ax.plot(x, y)
    ax.set_ylabel('Fitness')
    ax.set_xlabel('Generation')
    ax.set_title('Knapsack Genetic Algorithm')

    ax.text(0.02, 0.95, f"Generation: {generation}\nFitness: {fitness}", 
            transform=ax.transAxes, fontsize=10, verticalalignment='top', bbox=dict(boxstyle="round", facecolor="wheat"))

    plt.tight_layout()

    canvas = FigureCanvasAgg(fig)
    canvas.draw()
    renderer = canvas.get_renderer()
    raw_data = renderer.tostring_argb()

    size = canvas.get_width_height()
    surf = pygame.image.fromstring(raw_data, size, "ARGB")
    screen.blit(surf, (0, 0))
    pygame.display.flip()

# Initialize Pygame
pygame.init()
screen = pygame.display.set_mode((600, 500))
pygame.display.set_caption("Genetic Algorithm - Knapsack Problem")

num_items = 30  # Number of items
max_weight = 10  # Maximum weight of an item
max_value = 20  # Maximum value of an item
max_capacity = 150  # Maximum capacity of the knapsack

# Generate random items and knapsack capacity
items, capacity = randomize_items_and_capacity(num_items, max_weight, max_value, max_capacity)

# Algorithm parameters
population_size = 20  # Population size
generations = 300  # Maximum number of generations
mutation_rate = 0.05  # Mutation rate

# Running the genetic algorithm
genetic_algorithm(population_size, num_items, generations, mutation_rate, items, capacity, screen)

# Close Pygame
pygame.quit()
